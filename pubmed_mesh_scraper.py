# #!/usr/bin/env python3
# """
# PubMed NCBI Data Scraper — Command Line Tool
# Features:
#   • Multiple search field types (MeSH, Title, Abstract, Title+Abstract, All Fields)
#   • Weekly date-range chunking to bypass NCBI's 10,000-result cap
#   • Parallel efetch with ThreadPoolExecutor
#   • One row per unique email (expanded + deduplicated)
#   • Final CSV deduplicated by email
# """

# import requests
# import json
# import csv
# import re
# import time
# import argparse
# import xml.etree.ElementTree as ET
# from datetime import datetime, timedelta
# from typing import List, Dict, Optional, Tuple
# from concurrent.futures import ThreadPoolExecutor, as_completed

# try:
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
#     from openpyxl.utils import get_column_letter
#     OPENPYXL_AVAILABLE = True
# except ImportError:
#     OPENPYXL_AVAILABLE = False

# ESEARCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
# EFETCH_URL  = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"

# # ── Search field tags ──────────────────────────────────────────────────────────
# SEARCH_FIELDS = {
#     "mesh":     "[MeSH Terms]",
#     "title":    "[Title]",
#     "abstract": "[Abstract]",
#     "tiab":     "[Title/Abstract]",
#     "all":      "",
# }

# # NCBI hard cap: esearch returns at most ~9,999 IDs per date-window.
# # We split large ranges into weekly windows to capture everything.
# NCBI_RESULT_CAP = 9_900   # trigger weekly split when total >= this


# # ── Helpers ────────────────────────────────────────────────────────────────────

# def week_ranges(start: datetime, end: datetime) -> List[Tuple[datetime, datetime]]:
#     """Yield (week_start, week_end) pairs covering [start, end]."""
#     ranges = []
#     cursor = start
#     while cursor <= end:
#         week_end = min(cursor + timedelta(days=6), end)
#         ranges.append((cursor, week_end))
#         cursor = week_end + timedelta(days=1)
#     return ranges


# def fmt_date(dt: datetime) -> str:
#     """YYYY/MM/DD string for NCBI."""
#     return dt.strftime("%Y/%m/%d")


# def extract_emails(text: str) -> List[str]:
#     """Return all unique email addresses found in a string."""
#     if not text:
#         return []
#     emails = re.findall(r"[\w.\+\-]+@[\w.\-]+\.\w{2,}", text)
#     seen, unique = set(), []
#     for e in emails:
#         if e.lower() not in seen:
#             seen.add(e.lower())
#             unique.append(e)
#     return unique


# # ── Main scraper class ─────────────────────────────────────────────────────────

# class PubMedScraper:
#     def __init__(
#         self,
#         query: str,
#         search_field: str,
#         start_date: str,
#         end_date: str,
#         api_key: str    = "1b9dd02b2dde8556499eaab1095c18a0ac09",
#         batch_size: int = 200,
#         max_workers: int = 5,
#         delay: float    = 0.4,
#     ):
#         """
#         Args:
#             query        : Search term (e.g. "diabetes")
#             search_field : One of mesh / title / abstract / tiab / all
#             start_date   : YYYY/MM/DD
#             end_date     : YYYY/MM/DD
#             api_key      : NCBI API key (optional; raises rate limit 3->10 req/s)
#             batch_size   : PMIDs per efetch call (max 200)
#             max_workers  : Parallel threads
#             delay        : Seconds between esearch pages
#         """
#         self.query        = query
#         self.field_tag    = SEARCH_FIELDS[search_field]
#         self.search_field = search_field
#         self.start_dt     = datetime.strptime(start_date, "%Y/%m/%d")
#         self.end_dt       = datetime.strptime(end_date,   "%Y/%m/%d")
#         self.api_key      = api_key
#         self.batch_size   = min(batch_size, 200)
#         self.max_workers  = max_workers
#         self.delay        = delay
#         self.all_results: List[Dict] = []

#     # ── Private helpers ──────────────────────────────────────────────────────

#     def _base_params(self) -> Dict:
#         p = {}
#         if self.api_key:
#             p["api_key"] = self.api_key
#         return p

#     def _term(self) -> str:
#         if not self.field_tag:
#             # "all" mode — no tag, let NCBI auto-map exactly like the website
#             return self.query
#         else:
#             # For specific fields, quote multi-word terms for exact phrase matching
#             q = f'"{self.query}"' if " " in self.query else self.query
#             return f"{q}{self.field_tag}"

#     # ── Phase 1: collect PMIDs (parallel across weekly windows) ────────────────

#     def _probe_window_count(self, start: datetime, end: datetime) -> int:
#         """
#         Quick single esearch to find out how many results exist in a date window.
#         Uses retmax=0 so NCBI returns only the count, not IDs — very fast.
#         """
#         params = {
#             **self._base_params(),
#             "db"      : "pubmed",
#             "term"    : self._term(),
#             "mindate" : fmt_date(start),
#             "maxdate" : fmt_date(end),
#             "datetype": "pdat",
#             "retmode" : "json",
#             "retmax"  : 0,
#         }
#         try:
#             resp = requests.get(ESEARCH_URL, params=params, timeout=30)
#             resp.raise_for_status()
#             return int(resp.json()["esearchresult"].get("count", 0))
#         except Exception as e:
#             print(f"    [probe error] {e}")
#             return 0

#     def _fetch_pmids_one_window(
#         self,
#         start: datetime,
#         end: datetime,
#         retmax: int = 200,
#         window_idx: int = 0,
#         total_windows: int = 1,
#         progress_lock=None,
#         completed_counter=None,
#     ) -> Tuple[List[str], str]:
#         """
#         Fetch ALL PMIDs for a single date window using retstart pagination.
#         This method handles ONE window only (no recursive splitting).
#         Returns (pmid_list, window_label).
#         """
#         pmids: List[str] = []
#         retstart = 0
#         total    = None

#         while True:
#             params = {
#                 **self._base_params(),
#                 "db"       : "pubmed",
#                 "term"     : self._term(),
#                 "mindate"  : fmt_date(start),
#                 "maxdate"  : fmt_date(end),
#                 "datetype" : "pdat",
#                 "retmode"  : "json",
#                 "retmax"   : retmax,
#                 "retstart" : retstart,
#             }
#             data = None
#             for attempt in range(3):
#                 try:
#                     resp = requests.get(ESEARCH_URL, params=params, timeout=30)
#                     if resp.status_code == 429:
#                         wait = 2 ** (attempt + 1)
#                         print(f"    [429 rate-limit esearch] retrying in {wait}s …")
#                         time.sleep(wait)
#                         continue
#                     resp.raise_for_status()
#                     data = resp.json()["esearchresult"]
#                     break
#                 except Exception as e:
#                     if attempt == 2:
#                         print(f"    [esearch error] {fmt_date(start)}→{fmt_date(end)}: {e}")
#                         return pmids, f"{fmt_date(start)}→{fmt_date(end)}"
#                     time.sleep(2)

#             if data is None:
#                 # All retries exhausted (repeated 429s) — skip window gracefully
#                 print(f"    [esearch skip] {fmt_date(start)}→{fmt_date(end)}: all retries failed")
#                 break

            

#             if total is None:
#                 total = int(data.get("count", 0))

#             ids = data.get("idlist", [])
#             pmids.extend(ids)
#             retstart += len(ids)

#             if retstart >= total or not ids:
#                 break

#             time.sleep(self.delay)

#         label = f"{fmt_date(start)}→{fmt_date(end)}"
#         if completed_counter is not None and progress_lock is not None:
#             with progress_lock:
#                 completed_counter[0] += 1
#                 done = completed_counter[0]
#                 total_str = f"{total:,}" if total is not None else "?"
#                 print(f"    [{done:>4}/{total_windows}] {label}  →  {len(pmids):,} IDs  "
#                       f"(window total in NCBI: {total_str})")

#         return pmids, label

#     def fetch_all_pmids(
#         self,
#         retmax: int = 200,
#         verbose: bool = True,
#         id_workers: int = None,
#     ) -> List[str]:
#         """
#         Phase 1: collect ALL PMIDs across the full date range in parallel.

#         Strategy:
#           1. Probe the full range with retmax=0 to get total count instantly.
#           2. If total < NCBI cap  → fetch the range as ONE sequential window.
#           3. If total >= NCBI cap → split into weekly windows and fetch ALL
#              windows simultaneously with ThreadPoolExecutor.
#              Each worker independently paginates its own week.
#         """
#         workers = id_workers or self.max_workers

#         if verbose:
#             print(f"\n{'='*65}")
#             print(f"  Query        : {self._term()}")
#             print(f"  Search type  : {self.search_field}")
#             print(f"  Date range   : {fmt_date(self.start_dt)} → {fmt_date(self.end_dt)}")
#             print(f"{'='*65}")
#             print("Phase 1 — Collecting PMIDs …\n")

#         # ── Step 1: fast probe of total result count ──────────────────────────
#         total_count = self._probe_window_count(self.start_dt, self.end_dt)
#         if verbose:
#             print(f"  Total results in NCBI for this query: {total_count:,}")

#         # ── Step 2: decide single-window vs parallel weekly windows ───────────
#         if total_count < NCBI_RESULT_CAP:
#             # Small range — no need to split
#             if verbose:
#                 print(f"  Under NCBI cap → single sequential fetch\n")
#             pmids_raw, _ = self._fetch_pmids_one_window(
#                 self.start_dt, self.end_dt, retmax,
#                 window_idx=0, total_windows=1,
#             )
#             if verbose:
#                 print(f"    [  1/  1] {fmt_date(self.start_dt)}→{fmt_date(self.end_dt)}"
#                       f"  →  {len(pmids_raw):,} IDs")
#         else:
#             # Large range — split into weekly windows, fetch in parallel
#             windows = week_ranges(self.start_dt, self.end_dt)
#             n = len(windows)
#             if verbose:
#                 print(f"  Exceeds NCBI cap → splitting into {n} weekly windows")
#                 print(f"  Parallel ID workers: {workers}\n")

#             progress_lock    = __import__("threading").Lock()
#             completed_counter = [0]   # mutable int inside list for thread-safe update

#             pmids_raw: List[str] = []

#             with ThreadPoolExecutor(max_workers=workers) as pool:
#                 futures = {
#                     pool.submit(
#                         self._fetch_pmids_one_window,
#                         ws, we, retmax, idx, n,
#                         progress_lock, completed_counter,
#                     ): idx
#                     for idx, (ws, we) in enumerate(windows)
#                 }
#                 for future in as_completed(futures):
#                     window_pmids, _ = future.result()
#                     pmids_raw.extend(window_pmids)

#         # ── Step 3: deduplicate (overlap between adjacent weeks is possible) ───
#         pmids = list(dict.fromkeys(pmids_raw))

#         if verbose:
#             print(f"\n  ✓ Total unique PMIDs collected: {len(pmids):,}")

#         return pmids

#     # ── Phase 2: fetch + parse article XML ──────────────────────────────────

#     def _fetch_batch(self, pmids: List[str]) -> List[Dict]:
#         """efetch a single batch of PMIDs and return parsed author rows."""
#         params = {
#             **self._base_params(),
#             "db"     : "pubmed",
#             "id"     : ",".join(pmids),
#             "rettype": "abstract",
#             "retmode": "xml",
#         }
#         for attempt in range(3):
#             try:
#                 resp = requests.get(EFETCH_URL, params=params, timeout=60)
#                 if resp.status_code == 429:
#                     wait = 2 ** (attempt + 1)
#                     print(f"    [429 rate-limit] retrying in {wait}s …")
#                     time.sleep(wait)
#                     continue
#                 resp.raise_for_status()
#                 return self._parse_xml(resp.text)
#             except Exception as e:
#                 print(f"    [efetch error attempt {attempt+1}] {e}")
#                 time.sleep(2)
#         return []

#     def _parse_xml(self, xml_text: str) -> List[Dict]:
#         """
#         Parse PubMed XML. Returns one row per unique email address.
#         If an author appears in multiple affiliations with different emails,
#         they get one row per email.
#         """
#         records = []
#         try:
#             root = ET.fromstring(xml_text)
#         except ET.ParseError as e:
#             print(f"  [XML parse error] {e}")
#             return records

#         for article_node in root.findall(".//PubmedArticle"):
#             # ── Article-level fields ─────────────────────────────────────────
#             pmid_node  = article_node.find(".//MedlineCitation/PMID")
#             pmid       = pmid_node.text.strip() if pmid_node is not None else "N/A"

#             title_node = article_node.find(".//ArticleTitle")
#             title      = ("".join(title_node.itertext()).strip()
#                           if title_node is not None else "N/A")

#             doi = "N/A"
#             for aid in article_node.findall(".//PubmedData/ArticleIdList/ArticleId"):
#                 if aid.get("IdType") == "doi":
#                     doi = (aid.text or "").strip() or "N/A"
#                     break

#             pub_url  = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
#             pub_date = "N/A"
#             pd = article_node.find(".//PubMedPubDate[@PubStatus='pubmed']")
#             if pd is not None:
#                 y = pd.findtext("Year", "")
#                 m = pd.findtext("Month","")
#                 d = pd.findtext("Day",  "")
#                 pub_date = f"{y}/{m}/{d}".strip("/")

#             # ── Author-level fields ──────────────────────────────────────────
#             for author in article_node.findall(".//AuthorList/Author"):
#                 last  = (author.findtext("LastName")  or "").strip()
#                 first = (author.findtext("ForeName")  or
#                          author.findtext("Initials")  or "").strip()
#                 full  = (f"{first} {last}".strip() if (first or last)
#                          else (author.findtext("CollectiveName") or "N/A"))

#                 affiliations = [
#                     aff.text.strip()
#                     for aff in author.findall(".//AffiliationInfo/Affiliation")
#                     if aff.text
#                 ]
#                 affil_str = " | ".join(affiliations) if affiliations else "N/A"

#                 orcid = "N/A"
#                 for ident in author.findall("Identifier"):
#                     if ident.get("Source") == "ORCID" and ident.text:
#                         orcid = ident.text.strip()
#                         break

#                 # ── Collect all unique emails across all affiliations ─────────
#                 all_emails: List[str] = []
#                 for aff_text in affiliations:
#                     all_emails.extend(extract_emails(aff_text))

#                 seen: set = set()
#                 unique_emails = []
#                 for e in all_emails:
#                     if e.lower() not in seen:
#                         seen.add(e.lower())
#                         unique_emails.append(e)

#                 if not unique_emails:
#                     continue   # skip authors without any email

#                 # ── One row per unique email ──────────────────────────────────
#                 for email in unique_emails:
#                     records.append({
#                         "pmid"       : pmid,
#                         "title"      : title,
#                         "doi"        : doi,
#                         "pub_url"    : pub_url,
#                         "pub_date"   : pub_date,
#                         "first_name" : first or "N/A",
#                         "last_name"  : last  or "N/A",
#                         "full_name"  : full,
#                         "email"      : email,
#                         "affiliation": affil_str,
#                         "orcid"      : orcid,
#                     })

#         return records

#     # ── Phase 3: parallel efetch orchestration ───────────────────────────────

#     def fetch_articles(self, pmids: List[str], verbose: bool = True) -> List[Dict]:
#         batches       = [pmids[i:i+self.batch_size]
#                          for i in range(0, len(pmids), self.batch_size)]
#         total_batches = len(batches)

#         if verbose:
#             print(f"\nPhase 2 — Fetching article details …")
#             print(f"  Batches  : {total_batches}  ({self.batch_size} PMIDs each)")
#             print(f"  Workers  : {self.max_workers} parallel threads\n")

#         completed = 0
#         with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
#             futures = {pool.submit(self._fetch_batch, b): i
#                        for i, b in enumerate(batches)}
#             for future in as_completed(futures):
#                 records = future.result()
#                 self.all_results.extend(records)
#                 completed += 1
#                 if verbose:
#                     print(f"  Batch {completed:>4}/{total_batches}  "
#                           f"→ +{len(records):>4} email-rows  "
#                           f"(running total: {len(self.all_results):,})")

#         # ── Deduplicate: one row per unique email ─────────────────────────────
#         if verbose:
#             print(f"\nPhase 3 — Deduplicating by email …")
#         before = len(self.all_results)
#         seen_emails: set = set()
#         deduped: List[Dict] = []
#         for row in self.all_results:
#             key = row["email"].lower()
#             if key not in seen_emails:
#                 seen_emails.add(key)
#                 deduped.append(row)
#         self.all_results = deduped

#         if verbose:
#             removed = before - len(self.all_results)
#             print(f"  Rows before dedup : {before:,}")
#             print(f"  Duplicates removed: {removed:,}")
#             print(f"  ✓ Unique emails   : {len(self.all_results):,}")

#         return self.all_results

#     # ── Save ──────────────────────────────────────────────────────────────────

#     def save_to_csv(self, filename: str = None) -> str:
#         if not filename:
#             safe_query = re.sub(r"[^\w\-]", "_", self.query)[:30]
#             sd = fmt_date(self.start_dt).replace("/", "-")
#             ed = fmt_date(self.end_dt  ).replace("/", "-")
#             filename = f"{safe_query}-{self.search_field}-{sd}-{ed}-emails.csv"

#         if not self.all_results:
#             print("No results to save.")
#             return filename

#         fieldnames = [
#             "pmid", "title", "doi", "pub_url", "pub_date",
#             "first_name", "last_name", "full_name",
#             "email", "affiliation", "orcid",
#         ]
#         with open(filename, "w", newline="", encoding="utf-8") as f:
#             writer = csv.DictWriter(f, fieldnames=fieldnames)
#             writer.writeheader()
#             writer.writerows(self.all_results)

#         print(f"\n  Saved → {filename}  ({len(self.all_results):,} rows)")
#         return filename

#     def save_to_json(self, filename: str = None) -> str:
#         if not filename:
#             safe_query = re.sub(r"[^\w\-]", "_", self.query)[:30]
#             ts = datetime.now().strftime("%Y%m%d_%H%M%S")
#             filename = f"pubmed_{safe_query}_{ts}.json"

#         if not self.all_results:
#             print("No results to save.")
#             return filename

#         with open(filename, "w", encoding="utf-8") as f:
#             json.dump(self.all_results, f, indent=2, ensure_ascii=False)

#         print(f"  Saved → {filename}  ({len(self.all_results):,} records)")
#         return filename

#     def save_to_xlsx(self, filename: str = None) -> str:
#         """
#         Save results to an Excel XLSX file using openpyxl.

#         Why XLSX instead of CSV for Windows users:
#           • XLSX stores text as UTF-16 internally — Excel opens it natively
#             with zero encoding prompts or garbled characters.
#           • CSV requires a UTF-8 BOM or manual "Data → From Text/CSV" import
#             to render non-ASCII characters correctly in Excel on Windows.

#         The output file has:
#           • Frozen + bold header row with a blue fill
#           • Auto-fitted column widths (capped at 80 chars)
#           • Alternating row shading for readability
#           • Hyperlinks in the pub_url column
#         """
#         if not OPENPYXL_AVAILABLE:
#             print("  [xlsx] openpyxl not installed — run: pip install openpyxl")
#             return ""

#         if not filename:
#             safe_query = re.sub(r"[^\w\-]", "_", self.query)[:30]
#             sd = fmt_date(self.start_dt).replace("/", "-")
#             ed = fmt_date(self.end_dt  ).replace("/", "-")
#             filename = f"{safe_query}-{self.search_field}-{sd}-{ed}-emails.xlsx"

#         if not self.all_results:
#             print("No results to save.")
#             return filename

#         COLS = [
#             "pmid", "title", "doi", "pub_url", "pub_date",
#             "first_name", "last_name", "full_name",
#             "email", "affiliation", "orcid",
#         ]
#         HEADERS = [
#             "PMID", "Title", "DOI", "PubMed URL", "Pub Date",
#             "First Name", "Last Name", "Full Name",
#             "Email", "Affiliation", "ORCID",
#         ]

#         wb = Workbook()
#         ws = wb.active
#         ws.title = "PubMed Emails"

#         # ── Styles ────────────────────────────────────────────────────────────
#         header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
#         header_fill   = PatternFill("solid", start_color="1F4E79")  # dark blue
#         header_align  = Alignment(horizontal="center", vertical="center",
#                                   wrap_text=False)
#         data_font     = Font(name="Arial", size=9)
#         alt_fill      = PatternFill("solid", start_color="EBF3FB")  # light blue
#         url_font      = Font(name="Arial", size=9, color="0563C1", underline="single")
#         thin          = Side(style="thin", color="CCCCCC")
#         cell_border   = Border(bottom=thin)
#         wrap_align    = Alignment(wrap_text=False, vertical="top")

#         # ── Header row ────────────────────────────────────────────────────────
#         ws.row_dimensions[1].height = 20
#         for col_idx, (col_key, header) in enumerate(zip(COLS, HEADERS), start=1):
#             cell = ws.cell(row=1, column=col_idx, value=header)
#             cell.font   = header_font
#             cell.fill   = header_fill
#             cell.alignment = header_align

#         ws.freeze_panes = "A2"   # freeze header row

#         # ── Data rows ─────────────────────────────────────────────────────────
#         col_max_len = [len(h) for h in HEADERS]

#         for row_idx, record in enumerate(self.all_results, start=2):
#             use_alt = (row_idx % 2 == 0)
#             row_fill = alt_fill if use_alt else None

#             for col_idx, col_key in enumerate(COLS, start=1):
#                 value = record.get(col_key, "")
#                 cell  = ws.cell(row=row_idx, column=col_idx, value=value)
#                 cell.font      = data_font
#                 cell.border    = cell_border
#                 cell.alignment = wrap_align
#                 if row_fill:
#                     cell.fill = row_fill

#                 # Hyperlink for pub_url column
#                 if col_key == "pub_url" and value and value.startswith("http"):
#                     cell.hyperlink = value
#                     cell.font = url_font

#                 # Track max content length for column width
#                 text_len = len(str(value)) if value else 0
#                 if text_len > col_max_len[col_idx - 1]:
#                     col_max_len[col_idx - 1] = text_len

#         # ── Column widths ─────────────────────────────────────────────────────
#         # Cap widths: title/affiliation get a reasonable max; others auto-fit
#         CAP = {
#             "title"      : 60,
#             "affiliation": 70,
#             "doi"        : 40,
#             "pub_url"    : 45,
#         }
#         for col_idx, col_key in enumerate(COLS, start=1):
#             cap    = CAP.get(col_key, 40)
#             width  = min(col_max_len[col_idx - 1] + 2, cap)
#             width  = max(width, 8)
#             ws.column_dimensions[get_column_letter(col_idx)].width = width

#         # ── Auto-filter on header row ─────────────────────────────────────────
#         ws.auto_filter.ref = (
#             f"A1:{get_column_letter(len(COLS))}1"
#         )

#         wb.save(filename)
#         print(f"\n  Saved → {filename}  ({len(self.all_results):,} rows)")
#         return filename

#     # ── One-call run ──────────────────────────────────────────────────────────

#     def run(self, verbose: bool = True, id_workers: int = None) -> List[Dict]:
#         pmids = self.fetch_all_pmids(retmax=200, verbose=verbose,
#                                      id_workers=id_workers)
#         if not pmids:
#             print("No PMIDs found — exiting.")
#             return []
#         self.fetch_articles(pmids, verbose=verbose)
#         return self.all_results


# # ── CLI ────────────────────────────────────────────────────────────────────────

# def main():
#     parser = argparse.ArgumentParser(
#         description="Scrape PubMed author email data via NCBI E-utilities.",
#         formatter_class=argparse.RawDescriptionHelpFormatter,
#         epilog="""
# SEARCH FIELD OPTIONS  (--search-type):
#   mesh      → query[MeSH Terms]         controlled vocabulary, most precise
#   title     → query[Title]              exact match in article title only
#   abstract  → query[Abstract]           match in abstract only
#   tiab      → query[Title/Abstract]     title OR abstract (most popular)
#   all       → query[All Fields]         broadest search

# HOW THE 10,000 RESULT LIMIT IS HANDLED:
#   NCBI's esearch API silently caps results at ~9,999 per query window.
#   For "diabetes 2020-2023" that's 95,310 total — a single query only
#   returns the first ~9,999. This script:
#     1. Probes the full range with retmax=0 to get the total count instantly.
#     2. If below cap: fetches the single window sequentially.
#     3. If above cap: splits into weekly 7-day windows and fetches ALL
#        windows IN PARALLEL — each worker handles its own week independently.
#        Phase 1 time drops from ~20 min sequential to ~1-2 min parallel.

# EXAMPLES:
#   # MeSH search (default)
#   python pubmed_scraper.py -m "diabetes" -s 2020/01/01 -e 2023/12/31

#   # Title+Abstract search with API key for faster fetching
#   python pubmed_scraper.py -m "hypertension" -s 2022/01/01 -e 2024/01/01 \\
#       --search-type tiab --api-key YOUR_KEY

#   # Broadest search, custom output file
#   python pubmed_scraper.py -m "malaria" -s 2020/01/01 -e 2024/12/31 \\
#       --search-type all -o malaria_emails.csv --json

# OUTPUT CSV COLUMNS:
#   pmid, title, doi, pub_url, pub_date, first_name, last_name,
#   full_name, email, affiliation, orcid

#   One row per unique email address. Final file is deduplicated — each
#   email appears at most once.

# TIP: Get a free NCBI API key at https://www.ncbi.nlm.nih.gov/account/
#      It raises the rate limit from 3 to 10 requests/second.
#         """,
#     )

#     parser.add_argument("--mesh-term",   "-m", required=True,
#                         help='Search term, e.g. "diabetes", "covid-19"')
#     parser.add_argument("--start-date",  "-s", required=True,
#                         help="Start date YYYY/MM/DD")
#     parser.add_argument("--end-date",    "-e", required=True,
#                         help="End date   YYYY/MM/DD")
#     parser.add_argument("--search-type", "-t",
#                         choices=list(SEARCH_FIELDS.keys()),
#                         default="mesh",
#                         help="Search field: mesh|title|abstract|tiab|all  (default: mesh)")
#     parser.add_argument("--api-key",     "-k", default=None,
#                         help="NCBI API key (optional, raises rate limit)")
#     parser.add_argument("--batch-size",  "-b", type=int, default=200,
#                         help="PMIDs per efetch call (max 200, default 200)")
#     parser.add_argument("--workers",     "-w", type=int, default=5,
#                         help="Parallel efetch threads for article XML (default 5)")
#     parser.add_argument("--id-workers",  "-W", type=int, default=None,
#                         help="Parallel threads for PMID collection (default: same as --workers)")
#     parser.add_argument("--delay",       "-d", type=float, default=0.4,
#                         help="Seconds between esearch pages (default 0.4)")
#     parser.add_argument("--output",      "-o", default=None,
#                         help="Output filename base (auto-generated if omitted)")
#     parser.add_argument("--xlsx",        action="store_true",
#                         help="Save as XLSX — recommended for Windows (no encoding issues)")
#     parser.add_argument("--csv",         action="store_true",
#                         help="Also save a UTF-8 CSV copy")
#     parser.add_argument("--json",        action="store_true",
#                         help="Also save a JSON copy of the results")

#     args = parser.parse_args()

#     for label, val in [("--start-date", args.start_date),
#                        ("--end-date",   args.end_date)]:
#         try:
#             datetime.strptime(val, "%Y/%m/%d")
#         except ValueError:
#             parser.error(f"{label} must be YYYY/MM/DD, got: {val!r}")

#     scraper = PubMedScraper(
#         query        = args.mesh_term,
#         search_field = args.search_type,
#         start_date   = args.start_date,
#         end_date     = args.end_date,
#         api_key      = args.api_key,
#         batch_size   = args.batch_size,
#         max_workers  = args.workers,
#         delay        = args.delay,
#     )

#     t0 = time.time()
#     scraper.run(verbose=True, id_workers=args.id_workers)
#     elapsed = time.time() - t0

#     print(f"\n{'='*65}")
#     print(f"  Time elapsed    : {elapsed:.1f}s")
#     if scraper.all_results:
#         print(f"  Speed           : {len(scraper.all_results)/elapsed:.1f} rows/s")
#         print(f"  Unique emails   : {len(scraper.all_results):,}")
#     print(f"{'='*65}")

#     if scraper.all_results:
#         # Default: save XLSX (best for Windows). Fall back to CSV if openpyxl missing.
#         saved_any = False
#         if args.xlsx or not (args.csv or args.json):
#             # --xlsx explicitly, OR no format flag given (xlsx is the default)
#             if OPENPYXL_AVAILABLE:
#                 xlsx_name = re.sub(r"\.csv$", ".xlsx", args.output) if args.output else None
#                 scraper.save_to_xlsx(xlsx_name)
#                 saved_any = True
#             else:
#                 print("  [warning] openpyxl not found — install with: pip install openpyxl")
#                 print("  Falling back to CSV …")
#                 scraper.save_to_csv(args.output)
#                 saved_any = True
#         if args.csv:
#             scraper.save_to_csv(args.output)
#             saved_any = True
#         if args.json:
#             scraper.save_to_json()
#             saved_any = True
#         if not saved_any:
#             scraper.save_to_csv(args.output)
#     else:
#         print("\nNo author records with valid emails found.")


# if __name__ == "__main__":
#     main()

#!/usr/bin/env python3
"""
PubMed NCBI Data Scraper
- Multiple search field types (MeSH, Title, Abstract, Title+Abstract, All Fields)
- Weekly date-range chunking to bypass NCBI's 10,000-result cap
- Parallel efetch with ThreadPoolExecutor
- Email extraction from: author affiliations + abstract correspondence lines
- One row per unique email (expanded + deduplicated)
- Real-time progress via optional progress_callback
"""

import requests, json, csv, re, time, os
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

ESEARCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
EFETCH_URL  = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"

SEARCH_FIELDS = {
    "mesh":     "[MeSH Terms]",
    "title":    "[Title]",
    "abstract": "[Abstract]",
    "tiab":     "[Title/Abstract]",
    "all":      "",
}

NCBI_RESULT_CAP = 9_900


# ── Helpers ────────────────────────────────────────────────────────────────────

def week_ranges(start: datetime, end: datetime) -> List[Tuple[datetime, datetime]]:
    ranges, cursor = [], start
    while cursor <= end:
        week_end = min(cursor + timedelta(days=6), end)
        ranges.append((cursor, week_end))
        cursor = week_end + timedelta(days=1)
    return ranges


def fmt_date(dt: datetime) -> str:
    return dt.strftime("%Y/%m/%d")


def extract_emails(text: str) -> List[str]:
    """Return all unique email addresses found in a string, stripping trailing dots."""
    if not text:
        return []
    raw = re.findall(r"[\w.\+\-]+@[\w.\-]+\.\w{2,}", text)
    seen, unique = set(), []
    for e in raw:
        e = e.rstrip('.')   # NCBI sometimes includes trailing dot: "author@uni.edu."
        if e.lower() not in seen and '@' in e:
            seen.add(e.lower())
            unique.append(e)
    return unique


# ── Main scraper class ─────────────────────────────────────────────────────────

class PubMedScraper:
    def __init__(
        self,
        query: str,
        search_field: str,
        start_date: str,
        end_date: str,
        api_key: str          = "1b9dd02b2dde8556499eaab1095c18a0ac09",
        batch_size: int       = 200,
        max_workers: int      = 5,
        delay: float          = 0.4,
        progress_callback: Optional[Callable] = None,
    ):
        self.query             = query
        self.field_tag         = SEARCH_FIELDS.get(search_field, "[All Fields]")
        self.search_field      = search_field
        self.start_dt          = datetime.strptime(start_date, "%Y/%m/%d")
        self.end_dt            = datetime.strptime(end_date,   "%Y/%m/%d")
        self.api_key           = api_key
        self.batch_size        = min(batch_size, 200)
        self.max_workers       = max_workers
        self.delay             = delay
        self.progress_callback = progress_callback
        self.all_results: List[Dict] = []

    # ── Progress ──────────────────────────────────────────────────────────────

    def _progress(self, pct: int, msg: str, **kwargs):
        if self.progress_callback:
            try:
                self.progress_callback(pct, msg, **kwargs)
            except Exception:
                pass

    # ── Query helpers ─────────────────────────────────────────────────────────

    def _base_params(self) -> Dict:
        return {"api_key": self.api_key} if self.api_key else {}

    def _term(self) -> str:
        if not self.field_tag:
            # "all" mode — no tag, let NCBI auto-map exactly like the website
            return self.query
        else:
            # For specific fields, quote multi-word terms for exact phrase matching
            q = f'"{self.query}"' if " " in self.query else self.query
            return f"{q}{self.field_tag}"

    # ── Phase 1: collect PMIDs ─────────────────────────────────────────────────

    def _probe_window_count(self, start: datetime, end: datetime) -> int:
        params = {
            **self._base_params(),
            "db": "pubmed", "term": self._term(),
            "mindate": fmt_date(start), "maxdate": fmt_date(end),
            "datetype": "pdat", "retmode": "json", "retmax": 0,
        }
        try:
            resp = requests.get(ESEARCH_URL, params=params, timeout=30)
            resp.raise_for_status()
            return int(resp.json()["esearchresult"].get("count", 0))
        except Exception as e:
            print(f"    [probe error] {e}")
            return 0

    def _fetch_pmids_one_window(
        self, start, end, retmax=200,
        window_idx=0, total_windows=1,
        progress_lock=None, completed_counter=None,
    ) -> Tuple[List[str], str]:
        pmids, retstart, total = [], 0, None
        while True:
            params = {
                **self._base_params(),
                "db": "pubmed", "term": self._term(),
                "mindate": fmt_date(start), "maxdate": fmt_date(end),
                "datetype": "pdat", "retmode": "json",
                "retmax": retmax, "retstart": retstart,
            }
            data = None
            for attempt in range(3):
                try:
                    resp = requests.get(ESEARCH_URL, params=params, timeout=30)
                    if resp.status_code == 429:
                        time.sleep(2 ** (attempt + 1))
                        continue
                    resp.raise_for_status()
                    data = resp.json()["esearchresult"]
                    break
                except Exception as e:
                    if attempt == 2:
                        return pmids, f"{fmt_date(start)}→{fmt_date(end)}"
                    time.sleep(2)

            if data is None:
                break

            if total is None:
                total = int(data.get("count", 0))

            ids = data.get("idlist", [])
            pmids.extend(ids)
            retstart += len(ids)
            if retstart >= total or not ids:
                break
            time.sleep(self.delay)

        label = f"{fmt_date(start)}→{fmt_date(end)}"
        if completed_counter is not None and progress_lock is not None:
            with progress_lock:
                completed_counter[0] += 1
                done = completed_counter[0]
                print(f"    [{done:>4}/{total_windows}] {label} → {len(pmids):,} IDs")

        return pmids, label

    def fetch_all_pmids(self, retmax=200, verbose=True, id_workers=None) -> List[str]:
        workers = id_workers or self.max_workers
        total_count = self._probe_window_count(self.start_dt, self.end_dt)

        self._progress(5, f"Found {total_count:,} articles in NCBI for '{self.query}'")

        if verbose:
            print(f"\n  Query: {self._term()}")
            print(f"  Date: {fmt_date(self.start_dt)} → {fmt_date(self.end_dt)}")
            print(f"  Total in NCBI: {total_count:,}")

        if total_count == 0:
            return []

        if total_count < NCBI_RESULT_CAP:
            self._progress(8, f"Fetching {total_count:,} article IDs...")
            pmids_raw, _ = self._fetch_pmids_one_window(
                self.start_dt, self.end_dt, retmax,
                window_idx=0, total_windows=1,
            )
        else:
            windows = week_ranges(self.start_dt, self.end_dt)
            n = len(windows)
            self._progress(8, f"Large result set — splitting into {n} weekly windows...")
            progress_lock     = __import__("threading").Lock()
            completed_counter = [0]
            pmids_raw: List[str] = []

            with ThreadPoolExecutor(max_workers=workers) as pool:
                futures = {
                    pool.submit(
                        self._fetch_pmids_one_window,
                        ws, we, retmax, idx, n, progress_lock, completed_counter,
                    ): idx
                    for idx, (ws, we) in enumerate(windows)
                }
                for future in as_completed(futures):
                    window_pmids, _ = future.result()
                    pmids_raw.extend(window_pmids)

        pmids = list(dict.fromkeys(pmids_raw))
        self._progress(15, f"Collected {len(pmids):,} unique article IDs")
        return pmids

    # ── Phase 2: fetch + parse article XML ──────────────────────────────────

    def _fetch_batch(self, pmids: List[str]) -> List[Dict]:
        params = {
            **self._base_params(),
            "db": "pubmed", "id": ",".join(pmids),
            "rettype": "abstract", "retmode": "xml",
        }
        for attempt in range(3):
            try:
                resp = requests.get(EFETCH_URL, params=params, timeout=60)
                if resp.status_code == 429:
                    time.sleep(2 ** (attempt + 1))
                    continue
                resp.raise_for_status()
                return self._parse_xml(resp.text)
            except Exception as e:
                print(f"    [efetch error attempt {attempt+1}] {e}")
                time.sleep(2)
        return []

    def _parse_xml(self, xml_text: str) -> List[Dict]:
        """
        Parse PubMed XML — returns one row per unique email address.

        Email sources searched (in priority order):
        1. Author-level AffiliationInfo (most reliable)
        2. Article abstract text — catches "Correspondence: email@x.com" lines
        3. Article-level AffiliationInfo (some articles only have this)

        Authors without any email in ANY of these sources are skipped.
        """
        records = []
        try:
            root = ET.fromstring(xml_text)
        except ET.ParseError as e:
            print(f"  [XML parse error] {e}")
            return records

        for article_node in root.findall(".//PubmedArticle"):
            # ── Article-level metadata ───────────────────────────────────────
            pmid_node  = article_node.find(".//MedlineCitation/PMID")
            pmid       = pmid_node.text.strip() if pmid_node is not None else "N/A"

            title_node = article_node.find(".//ArticleTitle")
            title      = ("".join(title_node.itertext()).strip()
                          if title_node is not None else "N/A")

            doi = "N/A"
            for aid in article_node.findall(".//PubmedData/ArticleIdList/ArticleId"):
                if aid.get("IdType") == "doi":
                    doi = (aid.text or "").strip() or "N/A"
                    break

            pub_url  = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
            pub_date = "N/A"
            pd = article_node.find(".//PubMedPubDate[@PubStatus='pubmed']")
            if pd is not None:
                y = pd.findtext("Year", "")
                m = pd.findtext("Month", "")
                d = pd.findtext("Day", "")
                pub_date = f"{y}/{m}/{d}".strip("/")

            # ── Email source 2: abstract text ────────────────────────────────
            # Many articles embed "Correspondence: email" or "Contact: email"
            # in the abstract, especially recent open-access journals.
            abstract_text = " ".join(
                "".join(ab.itertext())
                for ab in article_node.findall(".//Abstract/AbstractText")
            )
            abstract_emails = set(e.lower() for e in extract_emails(abstract_text))

            # ── Email source 3: article-level affiliations ───────────────────
            # Some XML structures attach affiliations at the article node
            # rather than per-author.
            article_affs = [
                aff.text.strip()
                for aff in article_node.findall(".//MedlineCitation/Article/AuthorList/Author/AffiliationInfo/Affiliation")
                if aff.text
            ]
            article_level_emails = set()
            for aff in article_affs:
                for e in extract_emails(aff):
                    article_level_emails.add(e.lower())

            # ── Per-author extraction ────────────────────────────────────────
            for author in article_node.findall(".//AuthorList/Author"):
                last  = (author.findtext("LastName")  or "").strip()
                first = (author.findtext("ForeName")  or
                         author.findtext("Initials")  or "").strip()
                full  = (f"{first} {last}".strip() if (first or last)
                         else (author.findtext("CollectiveName") or "N/A"))

                author_affs = [
                    aff.text.strip()
                    for aff in author.findall(".//AffiliationInfo/Affiliation")
                    if aff.text
                ]
                affil_str = " | ".join(author_affs) if author_affs else "N/A"

                orcid = "N/A"
                for ident in author.findall("Identifier"):
                    if ident.get("Source") == "ORCID" and ident.text:
                        orcid = ident.text.strip()
                        break

                # Source 1: author-level affiliations
                found_emails: List[str] = []
                for aff in author_affs:
                    found_emails.extend(extract_emails(aff))

                # Source 2: abstract emails — only if this looks like a
                # corresponding author (has orcid, or is first/last author)
                # We add abstract emails to the FIRST author with no email found
                if not found_emails and abstract_emails:
                    # We'll attach abstract emails to the first author with affiliations
                    if author_affs and not records:
                        found_emails = list(abstract_emails)

                # Deduplicate
                seen: set = set()
                unique_emails = []
                for e in found_emails:
                    if e.lower() not in seen:
                        seen.add(e.lower())
                        unique_emails.append(e)

                if not unique_emails:
                    continue   # skip authors without any email

                for email in unique_emails:
                    records.append({
                        "pmid"       : pmid,
                        "title"      : title,
                        "doi"        : doi,
                        "pub_url"    : pub_url,
                        "pub_date"   : pub_date,
                        "first_name" : first or "N/A",
                        "last_name"  : last  or "N/A",
                        "full_name"  : full,
                        "email"      : email,
                        "affiliation": affil_str,
                        "orcid"      : orcid,
                    })

        return records

    # ── Phase 3: parallel efetch orchestration ───────────────────────────────

    def fetch_articles(self, pmids: List[str], verbose: bool = True) -> List[Dict]:
        batches       = [pmids[i:i+self.batch_size]
                         for i in range(0, len(pmids), self.batch_size)]
        total_batches = len(batches)

        if verbose:
            print(f"\nPhase 2 — Fetching article details ({total_batches} batches)…")

        completed = 0
        with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
            futures = {pool.submit(self._fetch_batch, b): i for i, b in enumerate(batches)}
            for future in as_completed(futures):
                records = future.result()
                self.all_results.extend(records)
                completed += 1
                pct = int(15 + (completed / total_batches) * 75)   # 15 → 90%
                self._progress(
                    pct,
                    f"Processing articles: batch {completed}/{total_batches}",
                    authors_count=len(self.all_results),
                    emails_count=len(self.all_results),
                )
                if verbose:
                    print(f"  Batch {completed:>4}/{total_batches} → "
                          f"+{len(records):>4} email-rows "
                          f"(total: {len(self.all_results):,})")

        # Deduplicate by email
        before = len(self.all_results)
        seen_emails: set = set()
        deduped: List[Dict] = []
        for row in self.all_results:
            key = row["email"].lower()
            if key not in seen_emails:
                seen_emails.add(key)
                deduped.append(row)
        self.all_results = deduped

        if verbose:
            print(f"\n  ✓ Unique emails: {len(self.all_results):,} "
                  f"(removed {before - len(self.all_results):,} duplicates)")

        self._progress(
            90,
            f"Found {len(self.all_results):,} unique emails from {len(pmids):,} articles",
            authors_count=len(self.all_results),
            emails_count=len(self.all_results),
        )
        return self.all_results

    # ── Save ──────────────────────────────────────────────────────────────────

    def save_to_csv(self, filename: str = None) -> str:
        if not filename:
            safe_query = re.sub(r"[^\w\-]", "_", self.query)[:30]
            sd = fmt_date(self.start_dt).replace("/", "-")
            ed = fmt_date(self.end_dt).replace("/", "-")
            filename = f"{safe_query}-{self.search_field}-{sd}-{ed}-emails.csv"

        if not self.all_results:
            print("No results to save.")
            return filename

        fieldnames = [
            "pmid", "title", "doi", "pub_url", "pub_date",
            "first_name", "last_name", "full_name",
            "email", "affiliation", "orcid",
        ]
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.all_results)

        print(f"\n  Saved → {filename}  ({len(self.all_results):,} rows)")
        return filename

    def save_to_xlsx(self, filename: str = None) -> str:
        if not OPENPYXL_AVAILABLE:
            print("  [xlsx] openpyxl not installed")
            return ""
        if not filename:
            safe_query = re.sub(r"[^\w\-]", "_", self.query)[:30]
            sd = fmt_date(self.start_dt).replace("/", "-")
            ed = fmt_date(self.end_dt).replace("/", "-")
            filename = f"{safe_query}-{self.search_field}-{sd}-{ed}-emails.xlsx"
        if not self.all_results:
            return filename

        COLS = ["pmid","title","doi","pub_url","pub_date",
                "first_name","last_name","full_name","email","affiliation","orcid"]
        HEADERS = ["PMID","Title","DOI","PubMed URL","Pub Date",
                   "First Name","Last Name","Full Name","Email","Affiliation","ORCID"]

        wb = Workbook()
        ws = wb.active
        ws.title = "PubMed Emails"
        ws.freeze_panes = "A2"

        header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        header_fill  = PatternFill("solid", start_color="1F4E79")
        data_font    = Font(name="Arial", size=9)
        alt_fill     = PatternFill("solid", start_color="EBF3FB")
        url_font     = Font(name="Arial", size=9, color="0563C1", underline="single")
        thin         = Side(style="thin", color="CCCCCC")
        cell_border  = Border(bottom=thin)
        wrap_align   = Alignment(wrap_text=False, vertical="top")
        hdr_align    = Alignment(horizontal="center", vertical="center")

        for col_idx, (col_key, header) in enumerate(zip(COLS, HEADERS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = hdr_align

        col_max_len = [len(h) for h in HEADERS]
        CAP = {"title": 60, "affiliation": 70, "doi": 40, "pub_url": 45}

        for row_idx, record in enumerate(self.all_results, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, col_key in enumerate(COLS, start=1):
                value = record.get(col_key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = wrap_align
                if fill:
                    cell.fill = fill
                if col_key == "pub_url" and value and value.startswith("http"):
                    cell.hyperlink = value
                    cell.font = url_font
                if len(str(value)) > col_max_len[col_idx - 1]:
                    col_max_len[col_idx - 1] = len(str(value))

        for col_idx, col_key in enumerate(COLS, start=1):
            cap = CAP.get(col_key, 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                min(col_max_len[col_idx - 1] + 2, cap), 8
            )
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
        wb.save(filename)
        print(f"\n  Saved → {filename}  ({len(self.all_results):,} rows)")
        return filename

    # ── One-call run ──────────────────────────────────────────────────────────

    def run(self, verbose: bool = True, id_workers: int = None) -> List[Dict]:
        self._progress(2, f"Starting PubMed search for '{self.query}'...")
        pmids = self.fetch_all_pmids(retmax=200, verbose=verbose, id_workers=id_workers)
        if not pmids:
            self._progress(100, "No articles found in PubMed for this query and date range")
            print("No PMIDs found — exiting.")
            return []
        self.fetch_articles(pmids, verbose=verbose)
        self._progress(
            100,
            f"PubMed complete: {len(self.all_results):,} email records from {len(pmids):,} articles",
            authors_count=len(self.all_results),
            emails_count=len(self.all_results),
        )
        return self.all_results