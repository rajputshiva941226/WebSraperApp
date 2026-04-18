#!/usr/bin/env python3
"""
PDF Scraper — integrated with WebScraperApp ScraperAdapter.

Phase 1 — PDF extraction (no external API calls):
    Delegates to PDF_Scraper/pdf_extraction_module.PDFAuthorExtractor.
    Cascade: GROBID (if running on :8070) → PyMuPDF4LLM JSON → PyMuPDF blocks.
    Output CSV columns:
        Author_Name | Email | Affiliation | Page_Number | Extraction_Method | Article_URL

Phase 2 — API search + selenium email extraction:
    Reads Phase 1 CSV, deduplicates authors, then for each author:
      1. Searches PubMed / Semantic Scholar / OpenAlex for their papers.
      2. Selenium-scrapes each paper's journal page to extract emails.
    Uses IntegratedPaperSearchAndEmailExtractor from PDF_Scraper/.
    Output: integrated_results.csv with matched_author, matched_email,
            paper_title, paper_source, paper_year, all_authors, all_emails.

Called by ScraperAdapter as:
    PDFScraper(pdf_path=<path>, output_dir=..., job_id=...,
               conference_name=..., progress_callback=...)
"""

import os
import re
import csv
import sys
import time
import logging
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
import xml.etree.ElementTree as ET

logger = logging.getLogger(__name__)

# ── NCBI / EuropePMC constants ───────────────────────────────────────────────
_ESEARCH_URL  = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi'
_EFETCH_URL   = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi'
_EPMC_URL     = 'https://www.ebi.ac.uk/europepmc/webservices/rest/search'
_NCBI_API_KEY = '1b9dd02b2dde8556499eaab1095c18a0ac09'
_EMAIL_RE     = re.compile(r'[\w.+\-]+@[\w.\-]+\.\w{2,}')


# ─────────────────────────────────────────────────────────────────────────────
class PDFScraper:
    """Adapter-compatible PDF author + email extractor."""

    JOURNAL_NAME = 'PDF Document'

    def __init__(
        self,
        pdf_path: str,
        output_dir: str = 'results',
        job_id: Optional[str] = None,
        progress_callback=None,
        conference_name: str = 'default',
        grobid_server: str = 'http://localhost:8070',
        **kwargs,
    ):
        self.pdf_path    = pdf_path
        self.output_dir  = output_dir
        self.job_id      = job_id or 'pdf_job'
        self.progress_cb = progress_callback
        self.conference  = conference_name
        self.grobid_url  = grobid_server
        os.makedirs(output_dir, exist_ok=True)

    def set_progress_callback(self, cb):
        self.progress_cb = cb

    # ── Internal helpers ──────────────────────────────────────────────────────
    def _p(self, pct: int, msg: str, **kw):
        self._last_pct = pct
        if self.progress_cb:
            try:
                self.progress_cb(pct, msg, **kw)
            except KeyboardInterrupt:
                raise  # propagate cooperative stop signal
            except Exception:
                pass
        logger.info('[PDFScraper] %d%% — %s', pct, msg)

    def _stop_check(self):
        """Force a stop check without updating displayed progress."""
        pct = getattr(self, '_last_pct', 0)
        if self.progress_cb:
            try:
                self.progress_cb(pct, '')
            except KeyboardInterrupt:
                raise

    def _grobid_alive(self) -> bool:
        try:
            r = requests.get(f'{self.grobid_url}/api/isalive', timeout=3)
            return r.status_code == 200
        except Exception:
            return False

    def _ensure_grobid_running(self) -> bool:
        """
        Start the GROBID Docker container if it is not already reachable.
        Mirrors the DockerManager logic from Automated_pdf_extraction.py.
        Returns True when GROBID /api/isalive responds 200.
        """
        if self._grobid_alive():
            logger.info('[PDFScraper] GROBID already running at %s', self.grobid_url)
            return True

        logger.info('[PDFScraper] GROBID not reachable — attempting Docker start…')
        import subprocess

        GROBID_IMAGE   = 'lfoppiano/grobid:0.8.0'
        CONTAINER_NAME = 'grobid-server'
        port = self.grobid_url.rstrip('/').rsplit(':', 1)[-1]   # e.g. '8070'

        # Verify Docker is installed
        try:
            subprocess.run(['docker', '--version'],
                           capture_output=True, timeout=5, check=True)
        except Exception as exc:
            logger.warning('[PDFScraper] Docker not available: %s', exc)
            return False

        # Try to start an already-existing (stopped) container first
        start_result = subprocess.run(
            ['docker', 'start', CONTAINER_NAME],
            capture_output=True, text=True, timeout=15
        )
        if start_result.returncode != 0:
            # Container doesn't exist — create it from scratch
            logger.info('[PDFScraper] Creating new GROBID container on port %s…', port)
            run_result = subprocess.run(
                ['docker', 'run', '-d', '--init', '--ulimit', 'core=0',
                 '-p', f'{port}:8070', '--name', CONTAINER_NAME, GROBID_IMAGE],
                capture_output=True, text=True, timeout=60
            )
            if run_result.returncode != 0:
                logger.error('[PDFScraper] docker run failed: %s', run_result.stderr.strip())
                return False
            logger.info('[PDFScraper] docker run OK: %s', run_result.stdout.strip())
        else:
            logger.info('[PDFScraper] docker start OK (existing container resumed)')

        # Wait up to 120 s for GROBID to become healthy
        logger.info('[PDFScraper] Waiting for GROBID to initialise (up to 120 s)…')
        self._p(12, 'Waiting for GROBID Docker container to start…')
        for i in range(60):
            time.sleep(2)
            if self._grobid_alive():
                logger.info('[PDFScraper] GROBID ready after ~%d s', (i + 1) * 2)
                return True

        logger.warning('[PDFScraper] GROBID did not become ready within 120 s')
        return False

    def _warn_no_grobid(self):
        """Log and surface a user-visible warning when GROBID is not reachable."""
        msg = (
            f'GROBID not reachable at {self.grobid_url} '
            '(Docker not running?). Falling back to PyMuPDF for extraction.'
        )
        logger.warning('[PDFScraper] %s', msg)
        return msg

    # ── API-based email lookup helpers ────────────────────────────────────────
    def _ss_resolve_author(self, author: str, affil: str) -> Optional[str]:
        """Return Semantic Scholar authorId or None."""
        try:
            r = requests.get(
                'https://api.semanticscholar.org/graph/v1/author/search',
                params={'query': author, 'fields': 'authorId,name,affiliations', 'limit': 10},
                timeout=10,
            )
            if r.status_code != 200:
                return None
            data = r.json().get('data', [])
            if not data:
                return None
            # Prefer candidate whose affiliation matches
            clean_affil = affil.lower()[:60] if affil else ''
            for cand in data:
                if clean_affil:
                    for ca in cand.get('affiliations', []):
                        if any(w in ca.lower() for w in clean_affil.split() if len(w) > 3):
                            return cand['authorId']
            return data[0]['authorId']
        except Exception:
            return None

    def _ss_dois_for_author(self, author_id: str) -> List[str]:
        """Return up to 5 DOIs from Semantic Scholar for this author."""
        try:
            r = requests.get(
                f'https://api.semanticscholar.org/graph/v1/author/{author_id}/papers',
                params={'fields': 'externalIds', 'limit': 20},
                timeout=10,
            )
            if r.status_code != 200:
                return []
            dois = []
            for paper in r.json().get('data', []):
                doi = (paper.get('externalIds') or {}).get('DOI', '')
                if doi:
                    dois.append(doi)
                if len(dois) >= 5:
                    break
            return dois
        except Exception:
            return []

    def _crossref_email(self, doi: str, author: str) -> str:
        """Try to extract an email from Crossref metadata for a given DOI."""
        try:
            r = requests.get(
                f'https://api.crossref.org/works/{doi}',
                timeout=8,
                headers={'User-Agent': 'WebScraperApp/1.0 (mailto:admin@example.com)'},
            )
            if r.status_code != 200:
                return ''
            last = author.split()[-1].lower() if author else ''
            for a in r.json().get('message', {}).get('author', []):
                full_lower = f"{a.get('given','')} {a.get('family','')}".lower()
                if last and last in full_lower:
                    email = a.get('email', '')
                    if email and '@' in email:
                        return email.strip().lower()
        except Exception:
            pass
        return ''

    def _openalex_dois(self, author: str) -> List[str]:
        """Return up to 5 DOIs from OpenAlex for this author."""
        try:
            r = requests.get(
                'https://api.openalex.org/authors',
                params={'search': author, 'per-page': 5, 'mailto': 'admin@example.com'},
                timeout=10,
            )
            if r.status_code != 200:
                return []
            results = r.json().get('results', [])
            if not results:
                return []
            aid = results[0].get('id', '').split('/')[-1]
            wr = requests.get(
                'https://api.openalex.org/works',
                params={'filter': f'authorships.author.id:{aid}',
                        'per-page': 10, 'mailto': 'admin@example.com'},
                timeout=10,
            )
            if wr.status_code != 200:
                return []
            dois = []
            for work in wr.json().get('results', []):
                doi = (work.get('doi') or '').replace('https://doi.org/', '')
                if doi:
                    dois.append(doi)
                if len(dois) >= 5:
                    break
            return dois
        except Exception:
            return []

    # ── Phase 2 API helpers (no Selenium) ─────────────────────────────────────

    @staticmethod
    def _extract_emails_text(text: str) -> List[str]:
        if not text:
            return []
        found = _EMAIL_RE.findall(text)
        seen, unique = set(), []
        for e in found:
            if e.lower() not in seen:
                seen.add(e.lower())
                unique.append(e.lower())
        return unique

    @staticmethod
    def _name_matches(searched: str, found: str) -> bool:
        """True if found name is plausibly the same person as searched."""
        s = searched.lower().split()
        f = found.lower().split()
        if not s or not f:
            return False
        # last-name overlap
        if s[-1] not in f[-1] and f[-1] not in s[-1]:
            return False
        # first-name initial overlap (if both have first name)
        if len(s) > 1 and len(f) > 1 and s[0] and f[0]:
            if s[0][0] != f[0][0]:
                return False
        return True

    def _pubmed_esearch_pmids(self, author_name: str, max_results: int = 5) -> List[str]:
        """Return up to max_results PMIDs for articles authored by author_name."""
        try:
            r = requests.get(_ESEARCH_URL, params={
                'db': 'pubmed',
                'term': f'"{author_name}"[Author]',
                'retmax': max_results,
                'retmode': 'json',
                'api_key': _NCBI_API_KEY,
            }, timeout=15)
            if r.status_code == 200:
                return r.json().get('esearchresult', {}).get('idlist', [])
        except Exception:
            pass
        return []

    def _pubmed_efetch_parse(self, pmids: List[str], author_searched: str) -> List[Dict]:
        """Efetch XML for PMIDs, parse email rows that match author_searched."""
        if not pmids:
            return []
        try:
            r = requests.get(_EFETCH_URL, params={
                'db': 'pubmed', 'id': ','.join(pmids),
                'retmode': 'xml', 'rettype': 'abstract',
                'api_key': _NCBI_API_KEY,
            }, timeout=60)
            if r.status_code != 200:
                return []
            root = ET.fromstring(r.text)
        except Exception as exc:
            logger.warning('[PDFScraper] efetch failed for %s: %s', author_searched, exc)
            return []

        rows: List[Dict] = []
        for article in root.findall('.//PubmedArticle'):
            pmid_nd  = article.find('.//PMID')
            pmid     = pmid_nd.text.strip() if pmid_nd is not None else 'N/A'
            title_nd = article.find('.//ArticleTitle')
            title    = ''.join(title_nd.itertext()).strip() if title_nd is not None else 'N/A'
            doi      = 'N/A'
            for aid in article.findall('.//PubmedData/ArticleIdList/ArticleId'):
                if aid.get('IdType') == 'doi':
                    doi = (aid.text or '').strip() or 'N/A'
                    break
            pub_url  = f'https://pubmed.ncbi.nlm.nih.gov/{pmid}/'
            abs_text = ' '.join(
                ''.join(ab.itertext())
                for ab in article.findall('.//Abstract/AbstractText')
            )
            abs_emails = self._extract_emails_text(abs_text)

            for author in article.findall('.//AuthorList/Author'):
                last  = (author.findtext('LastName') or '').strip()
                first = (author.findtext('ForeName') or
                         author.findtext('Initials') or '').strip()
                full  = f'{first} {last}'.strip() or (
                    author.findtext('CollectiveName') or 'N/A')
                if not self._name_matches(author_searched, full):
                    continue
                affs     = [a.text.strip()
                            for a in author.findall('.//AffiliationInfo/Affiliation')
                            if a.text]
                affil_str = ' | '.join(affs) if affs else 'N/A'
                emails: List[str] = []
                for aff in affs:
                    emails.extend(self._extract_emails_text(aff))
                if not emails:
                    emails = abs_emails[:]
                seen_e: set = set()
                unique_e = [e for e in emails if not (e in seen_e or seen_e.add(e))]
                if unique_e:
                    for email in unique_e:
                        rows.append({
                            'author_searched': author_searched,
                            'matched_author_name': full,
                            'email': email,
                            'affiliation': affil_str,
                            'doi': doi, 'pmid': pmid,
                            'pub_url': pub_url, 'title': title,
                            'source': 'pubmed',
                        })
                else:
                    rows.append({
                        'author_searched': author_searched,
                        'matched_author_name': full,
                        'email': '',
                        'affiliation': affil_str,
                        'doi': doi, 'pmid': pmid,
                        'pub_url': pub_url, 'title': title,
                        'source': 'pubmed',
                    })
        return rows

    def _europepmc_search(self, author_name: str, max_results: int = 5) -> List[Dict]:
        """Search EuropePMC for author articles; parse email from affiliations."""
        rows: List[Dict] = []
        try:
            r = requests.get(_EPMC_URL, params={
                'query': f'AUTH:"{author_name}"',
                'format': 'json', 'resulttype': 'core',
                'pageSize': max_results,
            }, timeout=15)
            if r.status_code != 200:
                return rows
            results = r.json().get('resultList', {}).get('result', [])
        except Exception as exc:
            logger.warning('[PDFScraper] EuropePMC search failed for %s: %s',
                           author_name, exc)
            return rows

        for art in results:
            pmid    = str(art.get('pmid', 'N/A') or 'N/A')
            doi     = str(art.get('doi',  'N/A') or 'N/A')
            title   = str(art.get('title', 'N/A') or 'N/A')
            pub_url = (f'https://pubmed.ncbi.nlm.nih.gov/{pmid}/'
                       if pmid != 'N/A'
                       else f'https://europepmc.org/article/MED/{art.get("id","")}')

            for a in (art.get('authorList') or {}).get('author', []):
                full = (f"{a.get('firstName','')} {a.get('lastName','')}".strip()
                        or a.get('collectiveName', 'N/A'))
                if not self._name_matches(author_name, full):
                    continue
                aff_detail = a.get('authorAffiliationDetailsList') or {}
                aff_list   = aff_detail.get('authorAffiliation', []) \
                             if isinstance(aff_detail, dict) else []
                affs       = [af.get('affiliation', '')
                              for af in aff_list if af.get('affiliation')]
                affil_str  = ' | '.join(affs) if affs else 'N/A'
                emails: List[str] = []
                for aff in affs:
                    emails.extend(self._extract_emails_text(aff))
                if a.get('email'):
                    emails.append(a['email'].lower())
                seen_e: set = set()
                unique_e = [e for e in emails if not (e in seen_e or seen_e.add(e))]
                base = {'author_searched': author_name,
                        'matched_author_name': full,
                        'affiliation': affil_str,
                        'doi': doi, 'pmid': pmid,
                        'pub_url': pub_url, 'title': title,
                        'source': 'europepmc'}
                if unique_e:
                    for email in unique_e:
                        rows.append({**base, 'email': email})
                else:
                    rows.append({**base, 'email': ''})
        return rows

    def _lookup_email(self, author: str, affil: str) -> str:
        """Try Semantic Scholar → OpenAlex → Crossref to find email."""
        return self._lookup_email_with_doi(author, affil)['email']

    def _lookup_email_with_doi(self, author: str, affil: str) -> Dict:
        """Same as _lookup_email but also returns the DOI that produced the hit."""
        empty = {'email': '', 'doi': ''}
        if not author or len(author) < 4:
            return empty

        # Semantic Scholar path
        try:
            aid = self._ss_resolve_author(author, affil)
            if aid:
                for doi in self._ss_dois_for_author(aid):
                    email = self._crossref_email(doi, author)
                    if email:
                        return {'email': email, 'doi': doi}
                    time.sleep(0.3)
        except Exception:
            pass

        # OpenAlex path
        try:
            for doi in self._openalex_dois(author):
                email = self._crossref_email(doi, author)
                if email:
                    return {'email': email, 'doi': doi}
                time.sleep(0.3)
        except Exception:
            pass

        return empty

    # ── Phase 1 : PDF extraction via PDFAuthorExtractor (no API calls) ─────────
    def run_phase1(self) -> Tuple[str, Dict]:
        """
        Delegate entirely to PDF_Scraper/pdf_extraction_module.PDFAuthorExtractor.
        Cascade: GROBID (if reachable) → PyMuPDF4LLM JSON → PyMuPDF blocks.
        NO email regex scanning — conference / brochure PDFs contain org emails,
        not author emails.
        Output columns: Author_Name | Email | Affiliation | Page_Number |
                        Extraction_Method | Article_URL
        """
        pdf = Path(self.pdf_path)
        self._p(5, f'Validating PDF: {pdf.name}')
        if not pdf.exists():
            raise FileNotFoundError(f'PDF not found: {self.pdf_path}')

        # ── Import extraction module ──
        scraper_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'PDF_Scraper')
        if scraper_dir not in sys.path:
            sys.path.insert(0, scraper_dir)
        try:
            import importlib
            import pdf_extraction_module as _pem
            importlib.reload(_pem)          # re-run module-level imports each job
            from pdf_extraction_module import ExtractionConfig, PDFAuthorExtractor
        except ImportError as exc:
            raise RuntimeError(f'pdf_extraction_module unavailable: {exc}') from exc

        self._p(10, 'Initialising extraction pipeline…')

        # Use persistent work dir so pages/ tei/ out/ survive for inspection
        workdir = Path(self.output_dir) / f'{self.job_id}_work'
        workdir.mkdir(parents=True, exist_ok=True)
        logger.info('[PDFScraper] Work directory: %s', workdir)

        if True:  # replaces 'with tempfile.TemporaryDirectory()' — keep indent
            tmpdir = str(workdir)
            log_path = workdir / 'pdf_extraction_debug.log'
            logger.info('[PDFScraper] Debug log: %s', log_path)
            config = ExtractionConfig(
                output_dir     = workdir / 'out',
                temp_pages_dir = workdir / 'pages',
                temp_tei_dir   = workdir / 'tei',
                grobid_server  = self.grobid_url,
                log_path       = log_path,
            )

            # Try to start GROBID Docker if it isn't already alive
            grobid_alive = self._ensure_grobid_running()
            grobid_url_for_extractor = None
            if grobid_alive:
                grobid_url_for_extractor = self.grobid_url
                self._p(15, f'GROBID available at {self.grobid_url}')
                logger.info('[PDFScraper] GROBID reachable, using direct HTTP')
            else:
                self._p(15, f'⚠ {self._warn_no_grobid()}')

            extractor = PDFAuthorExtractor(config=config, grobid_url=grobid_url_for_extractor)

            self._stop_check()
            self._p(20, f'Processing PDF pages…')

            try:
                total_found, method_stats = extractor.process_pdf(pdf)
            except Exception as exc:
                logger.error('[PDFScraper] Extraction failed: %s', exc)
                raise

            # Log what GROBID actually wrote so we can debug naming mismatches
            tei_files = list((workdir / 'tei').glob('*')) if (workdir / 'tei').exists() else []
            logger.info('[PDFScraper] TEI dir contents (%d files): %s',
                        len(tei_files), [f.name for f in tei_files[:20]])

            self._stop_check()
            self._p(85, 'Reading extracted results…')

            # Find output CSV written by PDFAuthorExtractor
            out_csv = config.output_dir / f'{pdf.stem}_authors.csv'
            if not out_csv.exists():
                csvs = list(config.output_dir.glob('*.csv'))
                out_csv = csvs[0] if csvs else None

            rows_out: List[Dict] = []
            seen_names: set = set()
            if out_csv and out_csv.exists():
                with open(out_csv, 'r', encoding='utf-8', errors='replace') as fh:
                    for row in csv.DictReader(fh):
                        name = (row.get('author_name') or '').strip()
                        if not name:
                            continue
                        name_key = name.lower()
                        if name_key in seen_names:
                            continue  # deduplicate across pages
                        seen_names.add(name_key)
                        rows_out.append({
                            'Author_Name':       name,
                            'Email':             '',
                            'Affiliation':       (row.get('affiliation') or '').strip(),
                            'Page_Number':       row.get('page_number', ''),
                            'Extraction_Method': row.get('extraction_method', ''),
                            'Article_URL':       row.get('original_pdf', pdf.name),
                        })

        # ── Write Phase 1 CSV ──
        self._p(92, 'Writing Phase 1 CSV…')
        safe_stem   = re.sub(r'[^\w\-]', '_', pdf.stem)[:50]   # cap at 50 chars
        output_file = os.path.join(
            self.output_dir,
            f'{self.job_id}_{safe_stem}_phase1.csv',
        )
        fieldnames = ['Author_Name', 'Email', 'Affiliation',
                      'Page_Number', 'Extraction_Method', 'Article_URL']

        with open(output_file, 'w', newline='', encoding='utf-8') as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows_out)

        total    = len(rows_out)
        no_email = total  # Phase 1 never has emails — that's Phase 2's job
        methods_str = ', '.join(f'{m}: {c}' for m, c in (method_stats or {}).items())
        self._p(100, f'Phase 1 done — {total} unique authors extracted'
                     + (f' ({methods_str})' if methods_str else '')
                     + f'; {no_email} need Phase 2 API email search')
        logger.info('[PDFScraper] Phase1 output: %s  rows=%d', output_file, total)

        return output_file, {
            'scraper':         'pdf_scraper',
            'phase':           1,
            'pdf_file':        pdf.name,
            'output_file':     output_file,
            'authors_found':   total,
            'emails_found':    0,
        }

    # ── Phase 2 : API-only email search (no Selenium / Chrome) ──────────────
    def run_phase2(self, phase1_csv: str) -> Tuple[str, Dict]:
        """
        Read Phase 1 CSV, then for each unique author run a cascade:
          1. PubMed esearch → efetch XML → parse emails from affiliation strings
          2. EuropePMC REST search → parse emails from affiliation strings
          3. Semantic Scholar → DOIs → CrossRef email (fallback)
          4. OpenAlex → DOIs → CrossRef email (fallback)
        Output: <job_id>_phase2_emails.csv  (unique by email)
                <job_id>_phase2_emails.xlsx (same, formatted)
        """
        p1 = Path(phase1_csv)
        self._p(5, f'Loading Phase 1 results: {p1.name}')
        if not p1.exists():
            raise FileNotFoundError(f'Phase 1 CSV not found: {phase1_csv}')

        with open(p1, 'r', encoding='utf-8', errors='replace') as fh:
            records = list(csv.DictReader(fh))
        logger.info('[PDFScraper] Phase2 loaded %d rows from %s', len(records), p1.name)

        # Deduplicate authors (case-insensitive)
        seen_names: set = set()
        unique_authors: List[Dict] = []
        for row in records:
            name = (row.get('Author_Name') or '').strip()
            if name and name.lower() not in seen_names:
                seen_names.add(name.lower())
                unique_authors.append({
                    'author_name': name,
                    'affiliation': (row.get('Affiliation') or '').strip(),
                })

        total = len(unique_authors)
        self._p(10, f'Phase 2: {total} unique authors → PubMed + EuropePMC API…')
        logger.info('[PDFScraper] Phase2: %d unique authors', total)

        _FIELDS = [
            'author_searched', 'matched_author_name', 'email', 'affiliation',
            'doi', 'pmid', 'pub_url', 'title', 'source',
        ]
        seen_emails:       set       = set()
        unique_email_rows: List[Dict] = []
        authors_with_email = 0
        emails_found       = 0

        for idx, info in enumerate(unique_authors):
            self._stop_check()
            pct    = 10 + int(idx / max(total, 1) * 85)
            author = info['author_name']
            affil  = info['affiliation']
            self._p(pct, f'[{idx + 1}/{total}] {author[:60]}…')
            logger.info('[PDFScraper] Phase2 author: %s', author)

            author_rows: List[Dict] = []

            # ── 1. PubMed esearch + efetch ──────────────────────────────────
            try:
                pmids = self._pubmed_esearch_pmids(author, max_results=5)
                if pmids:
                    author_rows.extend(self._pubmed_efetch_parse(pmids, author))
                time.sleep(0.4)
            except Exception as exc:
                logger.warning('[PDFScraper] PubMed failed for %s: %s', author, exc)

            # ── 2. EuropePMC (if no email yet) ──────────────────────────────
            if not any(r.get('email') for r in author_rows):
                try:
                    author_rows.extend(self._europepmc_search(author, max_results=5))
                    time.sleep(0.4)
                except Exception as exc:
                    logger.warning('[PDFScraper] EuropePMC failed for %s: %s', author, exc)

            # ── 3. Semantic Scholar → CrossRef ──────────────────────────────
            if not any(r.get('email') for r in author_rows):
                try:
                    aid = self._ss_resolve_author(author, affil)
                    if aid:
                        for doi in self._ss_dois_for_author(aid):
                            email = self._crossref_email(doi, author)
                            if email:
                                author_rows.append({
                                    'author_searched': author,
                                    'matched_author_name': author,
                                    'email': email, 'affiliation': affil,
                                    'doi': doi, 'pmid': 'N/A',
                                    'pub_url': f'https://doi.org/{doi}',
                                    'title': 'N/A',
                                    'source': 'semantic_scholar+crossref',
                                })
                                break
                            time.sleep(0.3)
                    time.sleep(0.4)
                except Exception as exc:
                    logger.warning('[PDFScraper] SS/CrossRef failed for %s: %s', author, exc)

            # ── 4. OpenAlex → CrossRef ───────────────────────────────────────
            if not any(r.get('email') for r in author_rows):
                try:
                    for doi in self._openalex_dois(author):
                        email = self._crossref_email(doi, author)
                        if email:
                            author_rows.append({
                                'author_searched': author,
                                'matched_author_name': author,
                                'email': email, 'affiliation': affil,
                                'doi': doi, 'pmid': 'N/A',
                                'pub_url': f'https://doi.org/{doi}',
                                'title': 'N/A',
                                'source': 'openalex+crossref',
                            })
                            break
                        time.sleep(0.3)
                    time.sleep(0.4)
                except Exception as exc:
                    logger.warning('[PDFScraper] OpenAlex/CrossRef failed for %s: %s', author, exc)

            # ── Collect unique emails ────────────────────────────────────────
            found_email = False
            for row in author_rows:
                email = (row.get('email') or '').strip().lower()
                if email and '@' in email and email not in seen_emails:
                    seen_emails.add(email)
                    unique_email_rows.append({f: row.get(f, '') for f in _FIELDS})
                    emails_found += 1
                    found_email = True
            if found_email:
                authors_with_email += 1

            time.sleep(0.5)

        # ── Write unique-email CSV ───────────────────────────────────────────
        self._p(96, 'Writing unique-emails output…')
        out_stem = f'{self.job_id}_phase2_emails'
        out_csv  = os.path.join(self.output_dir, f'{out_stem}.csv')
        with open(out_csv, 'w', newline='', encoding='utf-8') as fh:
            writer = csv.DictWriter(fh, fieldnames=_FIELDS, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(unique_email_rows)
        logger.info('[PDFScraper] Phase2 CSV: %s  rows=%d', out_csv, len(unique_email_rows))

        # ── Write XLSX ───────────────────────────────────────────────────────
        out_xlsx = ''
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws = wb.active
            ws.title = 'Phase 2 Emails'
            ws.freeze_panes = 'A2'
            _HEADERS = ['Author Searched', 'Matched Author', 'Email', 'Affiliation',
                        'DOI', 'PMID', 'PubMed URL', 'Title', 'Source']
            hf = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            hfill = PatternFill('solid', start_color='1F4E79')
            ws.append(_HEADERS)
            for cell in ws[1]:
                cell.font = hf
                cell.fill = hfill
            for row in unique_email_rows:
                ws.append([row.get(f, '') for f in _FIELDS])
            col_widths = [max(len(str(h)), 12) for h in _HEADERS]
            for row in unique_email_rows:
                for i, f in enumerate(_FIELDS):
                    col_widths[i] = min(max(col_widths[i], len(str(row.get(f, '')))), 70)
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w + 2
            ws.auto_filter.ref = f'A1:{get_column_letter(len(_HEADERS))}1'
            out_xlsx = os.path.join(self.output_dir, f'{out_stem}.xlsx')
            wb.save(out_xlsx)
            logger.info('[PDFScraper] Phase2 XLSX: %s', out_xlsx)
        except Exception as exc:
            logger.warning('[PDFScraper] XLSX write skipped: %s', exc)

        output_file = out_xlsx if (out_xlsx and os.path.exists(out_xlsx)) else out_csv

        self._p(100, f'Phase 2 done — {total} authors searched, '
                     f'{authors_with_email} with emails, {emails_found} unique emails')
        logger.info('[PDFScraper] Phase2 done: output=%s unique_emails=%d',
                    output_file, emails_found)

        return output_file, {
            'scraper':             'pdf_scraper_phase2',
            'phase':               2,
            'phase1_csv':          phase1_csv,
            'output_file':         output_file,
            'authors_processed':   total,
            'authors_with_emails': authors_with_email,
            'unique_emails':       emails_found,
        }

    # ── Backwards-compatible entry point ─────────────────────────────────────
    def run(self) -> Tuple[str, Dict]:
        """Runs Phase 1 only (extraction without API calls). Use run_phase2() separately."""
        return self.run_phase1()


# ─────────────────────────────────────────────────────────────────────────────
class PDFScraperPhase2:
    """
    Adapter-compatible wrapper for Phase 2 (API email lookup).
    The ScraperAdapter passes the Phase 1 CSV path via the `pdf_path` kwarg
    (which carries the `keyword` field from the Job row).
    """

    JOURNAL_NAME = 'PDF Email Search (Phase 2)'

    def __init__(
        self,
        pdf_path: str,            # receives the phase1 CSV path stored in Job.keyword
        output_dir: str = 'results',
        job_id: Optional[str] = None,
        progress_callback=None,
        conference_name: str = 'default',
        **kwargs,
    ):
        self._phase1_csv  = pdf_path
        self._output_dir  = output_dir
        self._job_id      = job_id or 'pdf_phase2'
        self._progress_cb = progress_callback
        self._conference  = conference_name

    def set_progress_callback(self, cb):
        self._progress_cb = cb

    def run(self) -> Tuple[str, Dict]:
        scraper = PDFScraper(
            pdf_path        = '',          # not used in phase 2
            output_dir      = self._output_dir,
            job_id          = self._job_id,
            progress_callback = self._progress_cb,
            conference_name = self._conference,
        )
        return scraper.run_phase2(self._phase1_csv)
